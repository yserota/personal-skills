"""Publishers for DCI output."""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Any


@dataclass
class PublishResult:
    target: str
    rows_written: int


UPSERT_KEY = ("writer_id", "period_start", "period_end")


def _rows_to_grid(rows: list[dict[str, Any]]) -> list[list[str]]:
    if not rows:
        return []
    headers = list(rows[0].keys())
    grid: list[list[str]] = [headers]
    for row in rows:
        grid.append(["" if row.get(h) is None else str(row.get(h)) for h in headers])
    return grid


def _normalize_record(record: dict[str, Any], headers: list[str]) -> dict[str, Any]:
    normalized: dict[str, Any] = {}
    for header in headers:
        value = record.get(header)
        normalized[header] = "" if value is None else value
    return normalized


def _upsert_by_key(existing: list[dict[str, Any]], incoming: list[dict[str, Any]]) -> list[dict[str, Any]]:
    if not incoming:
        return existing

    headers = list(incoming[0].keys())
    keyed: dict[tuple[str, str, str], dict[str, Any]] = {}
    for row in existing:
        normalized = _normalize_record(row, headers)
        key = tuple(str(normalized.get(k, "")) for k in UPSERT_KEY)
        keyed[key] = normalized

    for row in incoming:
        normalized = _normalize_record(row, headers)
        key = tuple(str(normalized.get(k, "")) for k in UPSERT_KEY)
        keyed[key] = normalized

    merged = list(keyed.values())
    merged.sort(key=lambda r: tuple(str(r.get(k, "")) for k in UPSERT_KEY))
    return merged


def publish_google_sheet(
    rows: list[dict[str, Any]],
    audit_rows: list[dict[str, Any]],
    sheet_id: str,
    worksheet_name: str,
    audit_worksheet_name: str,
    service_account_json_path: str,
) -> PublishResult:
    try:
        import gspread
    except ImportError as exc:  # pragma: no cover - dependency check
        raise RuntimeError("gspread is not installed. Run: pip install -e .") from exc

    if not Path(service_account_json_path).exists():
        raise FileNotFoundError(f"Service account key not found: {service_account_json_path}")

    gc = gspread.service_account(filename=service_account_json_path)
    spreadsheet = gc.open_by_key(sheet_id)

    ws = _ensure_worksheet(spreadsheet, worksheet_name)
    existing_rows = ws.get_all_records() if rows else []
    merged_rows = _upsert_by_key(existing_rows, rows)
    ws.clear()
    score_grid = _rows_to_grid(merged_rows)
    if score_grid:
        ws.update(score_grid, value_input_option="RAW")

    audit_ws = _ensure_worksheet(spreadsheet, audit_worksheet_name)
    existing_audit = audit_ws.get_all_records()
    merged_audit = existing_audit + audit_rows
    audit_ws.clear()
    audit_grid = _rows_to_grid(merged_audit)
    if audit_grid:
        audit_ws.update(audit_grid, value_input_option="RAW")

    return PublishResult(target="google_sheets", rows_written=len(merged_rows))


def _ensure_worksheet(spreadsheet: Any, worksheet_name: str) -> Any:
    try:
        return spreadsheet.worksheet(worksheet_name)
    except Exception:
        return spreadsheet.add_worksheet(title=worksheet_name, rows=2000, cols=50)


def publish_excel(
    tabs: dict[str, list[dict[str, Any]]],
    output_path: str | Path,
) -> PublishResult:
    """Write all dashboard tabs into a single .xlsx workbook, one sheet per tab.

    Args:
        tabs: mapping of sheet name → list of row dicts.
        output_path: destination .xlsx file path.
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill
    except ImportError as exc:
        raise RuntimeError("openpyxl is not installed. Run: pip install openpyxl") from exc

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove default blank sheet

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(fill_type="solid", fgColor="2F5496")

    total_written = 0
    for sheet_name, rows in tabs.items():
        ws = wb.create_sheet(title=sheet_name[:31])  # Excel sheet names max 31 chars
        if not rows:
            continue
        headers = list(rows[0].keys())
        ws.append(headers)
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
        for row in rows:
            ws.append([row.get(h, "") for h in headers])
        for col in ws.columns:
            max_len = max((len(str(cell.value or "")) for cell in col), default=0)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 50)
        total_written += len(rows)

    wb.save(output_path)
    return PublishResult(target="excel", rows_written=total_written)


def publish_dashboard_tabs(
    tabs: dict[str, list[dict[str, Any]]],
    sheet_id: str,
    service_account_json_path: str,
) -> PublishResult:
    """Upload each dashboard tab as a named worksheet, creating or overwriting as needed.

    Args:
        tabs: mapping of worksheet name → list of row dicts (as produced by build_dashboard_exports).
        sheet_id: Google Sheets spreadsheet ID (from the URL).
        service_account_json_path: path to the GCP service-account key JSON file.
    """
    try:
        import gspread
    except ImportError as exc:
        raise RuntimeError("gspread is not installed. Run: pip install -e .") from exc

    if not Path(service_account_json_path).exists():
        raise FileNotFoundError(f"Service account key not found: {service_account_json_path}")

    gc = gspread.service_account(filename=service_account_json_path)
    spreadsheet = gc.open_by_key(sheet_id)

    total_written = 0
    for tab_name, rows in tabs.items():
        ws = _ensure_worksheet(spreadsheet, tab_name)
        ws.clear()
        grid = _rows_to_grid(rows)
        if grid:
            ws.update(grid, value_input_option="USER_ENTERED")
        total_written += len(rows)

    return PublishResult(target="google_sheets_dashboard", rows_written=total_written)
